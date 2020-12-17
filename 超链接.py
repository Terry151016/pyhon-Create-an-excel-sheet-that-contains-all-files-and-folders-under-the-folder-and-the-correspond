import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color

def get_filename(dirname):
    
    column0 = dirname.count('\\')
    #创建工作簿    
    wb = Workbook()
    
    #创建sheet
    sheet1 = wb.create_sheet(title="Data")
      
    row = 1
    #循环获取根目录，根目录下的文件及文件夹
    for roots, dirs, files in os.walk(dirname):
        #print(roots)
        #计算不同层级文件的列数
        column = roots.count('\\')
        column = column - column0+1
        
        #获取文件夹名
        
        root_name = roots.split('\\')[-1]
        #生成文件夹的超链接
        rowB = '=hyperlink("' + roots  + '","' + root_name + '")'
        sheet1.cell(row=row,column=column).value=rowB
        #设置单元格的颜色
        ft = Font( color='FF0000')
        sheet1.cell(row=row,column=column).font = ft

        column = column + 1
        for file in files:
            #获取文件并将其设定为超链接
            if file[0] != '.' and '~'not in file:
                #print(file)
                row = row + 1
                rowB = '=hyperlink("' + roots + '\\' + file + '","' + file + '")'
                sheet1.cell(row=row,column = column).value=rowB
                
        row =row+1
        
     #删除创建工作簿自动生成的sheet
     #若要使用改sheet
     #可以通过以下实现
     #sheet1 =  wb.active 
     #sheet1.title = "自动生成的sheet"
    del wb['Sheet']   
    wb.save("get_file.xlsx")         


if __name__ == '__main__':
   #将文件夹路径作为参数传入
    get_filename(r'D:\0000数据库\11个人资料及作业\嵌入式\嵌入式系统课程设计')

